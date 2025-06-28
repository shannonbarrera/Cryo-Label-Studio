import csv
import re
from datetime import datetime, date
import openpyxl as xlsx

def get_data_list_csv(input_file_path, textboxformatinput, date_format=None):
    """
    Extracts label data from a CSV file using defined format string and header names.
    Skips rows where all values are empty or None.

    Args:
        input_file_path (str): Path to the CSV file.
        textboxformatinput (str): Format string describing column layout using header names.

    Returns:
        list: Extracted label data.
    """
    label_data_list_format = get_label_data_list_format(textboxformatinput)

    batchdata = []
    with open(input_file_path, 'r', encoding='utf-8') as file:
        csv_reader = list(csv.reader(file))
        columns_in_csv = csv_reader[0]
        indices_for_labeldata = [columns_in_csv.index(col) for col in label_data_list_format if col in columns_in_csv]

        for row in csv_reader[1:]:
            data = []
            for index in indices_for_labeldata:
                if index < len(row):
                    val = row[index]
                    if isinstance(val, str):
                        val = val.strip()
                        if val == "":
                            data.append(None)
                        elif date_format != "Leave as is":
                            data.append(try_parse_date(val))
                        else:
                            data.append(val)

                else:
                    data.append(None)

            if not all(val is None for val in data):
                batchdata.append(data)

    return batchdata


def get_data_list_xlsx(input_file_path, textboxformatinput, date_format=None):
    """
    Extracts label data from an Excel (.xlsx) file.

    Args:
        input_file_path (str): Path to the Excel file.
        textboxformatinput (str): Column layout format using headers.
        date_format (str or None): User-selected date format, or "Leave as is".

    Returns:
        list: Extracted label data (preserves datetime objects or raw strings).
    """
    workbook = xlsx.load_workbook(filename=input_file_path, read_only=True)
    sheet = workbook.active
    raw_info = extract_label_info(sheet, textboxformatinput)

    cleaned_info = []
    for row in raw_info[1:]:  # skip header
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append(None)
            elif isinstance(cell, str):
                stripped = cell.strip()
                if stripped == "":
                    cleaned_row.append(None)
                elif date_format != "Leave as is":
                    cleaned_row.append(try_parse_date(stripped))
                else:
                    cleaned_row.append(str(cell)) 
            elif isinstance(cell, (datetime, date)):
                if date_format == "Leave as is":
                    cleaned_row.append(str(cell))  
                else:
                    cleaned_row.append(cell)  
            else:
                cleaned_row.append(cell)
  


        if any(cell is not None for cell in cleaned_row):
            cleaned_info.append(cleaned_row)

    workbook.close()
    return cleaned_info




# Read the worksheet and format the label info into a list of lists
def extract_label_info(sheet, textboxformatinput):
    """
    Extracts label data from a worksheet using the defined coordinates and format.

    Args:
        sheet (Worksheet): Active Excel worksheet.
        tablecoords (list): Table boundaries.
        textboxformatinput (str): Format string defining column layout.

    Returns:
        list: List of label entries (each entry is a list).
    """
    label_data_list_format = get_label_data_list_format(textboxformatinput)
    columns_in_sheet = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    batchdata = []
    indices_for_labeldata = [columns_in_sheet.index(col) for col in label_data_list_format if col in columns_in_sheet]
    for row in sheet.iter_rows():
        extracted = []
        for idx in indices_for_labeldata:
            if idx < len(row):
                extracted.append(row[idx].value)
        if extracted:
            batchdata.append(extracted)
    return batchdata

def get_label_data_list_format(textboxformatinput):
    """
    Converts a format string with letter references (e.g., 'B\nD, C\nE') into column indices.

    Args:
        textboxformatinput (str): Format string where letters refer to column indices.

    Returns:
        list: List of zero-based column indices.
    """
    findlist = re.findall(r'{(.*?)}', textboxformatinput)
    return findlist

def try_parse_date(value):
    """
    Tries to parse a string into a date object using common formats.
    Returns a date if successful, or the original string if not.
    """
    if not isinstance(value, str):
        return value

    value = value.strip()
    if value == "":
        return None

    # Common formats you want to support
    date_formats = [
        "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y",
        "%Y/%m/%d", "%m-%d-%Y", "%d-%m-%Y",
        "%m/%d/%y", "%d/%m/%y",  # ✅ two-digit year
        "%m-%d-%y", "%d-%m-%y",  # ✅ two-digit year with dashes
        "%b %d, %Y", "%B %d, %Y"
    ]


    for fmt in date_formats:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    return value  # Return original if nothing matched


def remove_duplicate_labels(data_list):
    seen = set()
    unique_data = []
    for row in data_list:
        row_tuple = tuple(row)
        if row_tuple not in seen:
            seen.add(row_tuple)
            unique_data.append(row)
    return unique_data